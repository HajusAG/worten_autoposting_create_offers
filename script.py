"""
Worten Autoposting v3.0: Pending -> Active (MCM-only + blind-import strategy)

Daily Coolify scheduled task. Strategy mirrors Conforama/Adeo autoposting v3:

  1) GET /api/mcm/products/sources/status/export   (~1 req, ~10s)
     -> classify all pending products LIVE / NOT_LIVE / NOT_IN_MCM
  2) For LIVE: blind-import all of them via /api/offers/imports.
     Mirakl uses (shop_id, shop_sku) as offer key:
       - If offer exists -> updated (idempotent no-op for same price/qty)
       - If offer missing -> created
     Error report -> failed SKUs (validation errors, missing price etc.)
     accepted = uploaded - failed.
  3) Odoo: accepted -> Active (Worten website 117),
           failed -> New Task,
           NOT_LIVE -> stay in Pending (no per-task comment),
           NOT_IN_MCM -> New Task.

Worten specifics:
  - Multi-channel marketplace: offers use price[channel=WRT_PT_ONLINE] +
    price[channel=WRT_ES_ONLINE]; the plain "price" column carries the max.
  - Single Odoo website (117 = "Worten PT | HAJUS") covers both PT and ES.
  - Pricelists: Portugal=20, Spain=24.
  - No eco-contribution / producer-id columns (Worten does not collect them
    via the offer schema; mirrors the existing n8n workflow).

Credentials read from environment variables (Coolify-style).
"""

import requests
import io
import os
import sys
import time
import traceback
import openpyxl
from datetime import datetime
from typing import Any

WEBSITE_IDS = [117]
PENDING_WEBSITE_IDS = [117]

# Worten is a multi-channel marketplace (PT + ES under the same shop).
PRICELIST_TO_CHANNEL = {
    20: "WRT_PT_ONLINE",  # Portugal
    24: "WRT_ES_ONLINE",  # Spain
}

STAGE_NEW_TASK = 1
STAGE_ACTIVE = 3
STAGE_PENDING = 6

BITRIX_WEBHOOK = os.environ.get(
    "BITRIX_WEBHOOK",
    "https://boni.team/rest/107/x25sj3ejw9vo88b7",
)
ALERT_TITLE = "🚨 Worten Autoposting Pending→Active | Coolify"
ALERT_ACCOMPLICE_ID = int(os.environ.get("ALERT_ACCOMPLICE_ID", "584"))

MCM_MAX_RETRIES = 5


def send_alert(error_message: str, error_details: str = ""):
    now = datetime.utcnow().strftime("%d.%m.%Y %H:%M UTC")
    try:
        r = requests.post(f"{BITRIX_WEBHOOK}/tasks.task.list", json={
            "filter": {"TITLE": ALERT_TITLE, "GROUP_ID": 1, "!STATUS": 5},
            "select": ["ID"], "order": {"ID": "desc"}, "limit": 1,
        }, timeout=15)
        tasks = r.json().get("result", {}).get("tasks", [])

        if tasks:
            requests.post(f"{BITRIX_WEBHOOK}/task.commentitem.add", json={
                "taskId": tasks[0]["id"],
                "fields": {
                    "POST_MESSAGE": (
                        f"[B]Повторная ошибка — {now}[/B]\n\n"
                        f"[B]Ошибка:[/B] {error_message}\n\n"
                        f"[B]Детали:[/B]\n{error_details}"
                    ),
                    "AUTHOR_ID": 107,
                },
            }, timeout=15)
        else:
            desc = (
                f"[B]Алерт от автоматики[/B]\n\n"
                f"[B]Процесс:[/B] Worten Autoposting Pending→Active (v3)\n"
                f"[B]Источник:[/B] Coolify scheduled task\n"
                f"[B]Время:[/B] {now}\n\n"
                f"[B]Ошибка:[/B]\n[color=red]{error_message}[/color]\n\n"
                + (f"[B]Детали:[/B]\n{error_details}" if error_details else "")
            )
            requests.post(f"{BITRIX_WEBHOOK}/tasks.task.add", json={"fields": {
                "TITLE": ALERT_TITLE, "DESCRIPTION": desc,
                "RESPONSIBLE_ID": 107, "CREATED_BY": 107,
                "ACCOMPLICES": [ALERT_ACCOMPLICE_ID],
                "GROUP_ID": 1, "STAGE_ID": 705, "PRIORITY": 2,
            }}, timeout=15)
    except Exception as alert_err:
        print(f"Failed to send Bitrix24 alert: {alert_err}")


def odoo_rpc(odoo: dict, model: str, method: str, domain: list, fields: list | None = None, limit: int = 0) -> list:
    kwargs: dict[str, Any] = {}
    if fields:
        kwargs["fields"] = fields
    if limit:
        kwargs["limit"] = limit
    payload = {
        "jsonrpc": "2.0", "method": "call", "id": 1,
        "params": {"service": "object", "method": "execute_kw",
            "args": [odoo["db"], odoo["uid"], odoo["api_key"], model, method, [domain], kwargs]},
    }
    resp = requests.post(odoo["url"], json=payload, timeout=60)
    resp.raise_for_status()
    return resp.json().get("result", [])


def odoo_write(odoo: dict, model: str, ids: list[int], vals: dict) -> Any:
    payload = {
        "jsonrpc": "2.0", "method": "call", "id": 1,
        "params": {"service": "object", "method": "execute_kw",
            "args": [odoo["db"], odoo["uid"], odoo["api_key"], model, "write", [ids, vals]]},
    }
    resp = requests.post(odoo["url"], json=payload, timeout=60)
    resp.raise_for_status()
    return resp.json().get("result")


def odoo_create(odoo: dict, model: str, vals: dict) -> Any:
    payload = {
        "jsonrpc": "2.0", "method": "call", "id": 1,
        "params": {"service": "object", "method": "execute_kw",
            "args": [odoo["db"], odoo["uid"], odoo["api_key"], model, "create", [vals]]},
    }
    resp = requests.post(odoo["url"], json=payload, timeout=60)
    resp.raise_for_status()
    return resp.json().get("result")


def mirakl_fetch_mcm_status_dump(worten: dict) -> list[dict]:
    headers = {"Authorization": worten["api_key"]}
    params = {"shop_id": worten["shop_id"]}
    for attempt in range(MCM_MAX_RETRIES):
        resp = requests.get(
            f"{worten['base_url']}/api/mcm/products/sources/status/export",
            headers=headers, params=params, timeout=120,
        )
        if resp.status_code == 429:
            ra = int(resp.headers.get("Retry-After", 30))
            print(f"  MCM dump 429, wait {ra}s ({attempt + 1}/{MCM_MAX_RETRIES})")
            time.sleep(ra)
            continue
        resp.raise_for_status()
        data = resp.json()
        if isinstance(data, list):
            return data
        if isinstance(data, dict):
            for k in ("items", "products", "results", "body"):
                v = data.get(k)
                if isinstance(v, list):
                    return v
        return []
    raise Exception("MCM status dump: exhausted retries on 429")


def mirakl_import_offers(worten: dict, xlsx_bytes: bytes) -> int:
    headers = {"Authorization": worten["api_key"]}
    files = {"file": ("offers.xlsx", xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    data = {"import_mode": "NORMAL", "operator_format": "true"}
    params = {"shop_id": worten["shop_id"]}
    resp = requests.post(f"{worten['base_url']}/api/offers/imports", headers=headers, files=files, data=data, params=params, timeout=60)
    resp.raise_for_status()
    return resp.json()["import_id"]


def mirakl_check_import(worten: dict, import_id: int) -> dict:
    headers = {"Authorization": worten["api_key"]}
    resp = requests.get(f"{worten['base_url']}/api/offers/imports/{import_id}", headers=headers, timeout=30)
    resp.raise_for_status()
    return resp.json()


def mirakl_get_failed_skus(worten: dict, import_id: int) -> set:
    headers = {"Authorization": worten["api_key"]}
    resp = requests.get(f"{worten['base_url']}/api/offers/imports/{import_id}/error_report", headers=headers, timeout=30)
    resp.raise_for_status()
    failed = set()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        header_row = [str(cell.value or "").strip().lower() for cell in ws[1]]
        sku_col = None
        for idx, h in enumerate(header_row):
            if h in ("sku", "shop-sku", "shop_sku"):
                sku_col = idx
                break
        if sku_col is not None:
            for row in ws.iter_rows(min_row=2, values_only=True):
                val = str(row[sku_col] or "").strip()
                if val:
                    failed.add(val)
        wb.close()
    except Exception as e:
        print(f"  Error parsing XLSX error report: {e}")
    return failed


def clean_sku(sku: str) -> str:
    return sku.replace("BONI-", "", 1) if sku.startswith("BONI-") else sku


def build_offer_row(product: dict, prices_by_pricelist: dict[int, float]) -> dict | None:
    """Build multi-channel Worten offer row.

    Worten offers carry price[channel=WRT_PT_ONLINE] and price[channel=WRT_ES_ONLINE]
    plus a plain "price" fallback (set to max of channel prices). Rows without any
    valid channel price are dropped (caller sends task -> New Task).
    """
    sku = clean_sku(product.get("default_code", "") or "")
    if not sku:
        return None

    channel_prices: dict[str, float] = {}
    for pricelist_id, channel in PRICELIST_TO_CHANNEL.items():
        p = prices_by_pricelist.get(pricelist_id, 0) or 0
        if p > 0:
            channel_prices[channel] = p

    if not channel_prices:
        return None

    row: dict[str, Any] = {
        "sku": sku,
        "product-id": product.get("barcode", ""),
        "product-id-type": "EAN",
        "quantity": product.get("dr_free_qty", 0),
        "state": "11",
        "price": max(channel_prices.values()),
    }
    for channel, value in channel_prices.items():
        row[f"price[channel={channel}]"] = value
    return row


def build_xlsx(rows: list[dict]) -> bytes:
    if not rows:
        return b""
    all_keys = []
    seen = set()
    for r in rows:
        for k in r:
            if k not in seen:
                all_keys.append(k)
                seen.add(k)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(all_keys)
    for r in rows:
        ws.append([r.get(k, "") for k in all_keys])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def send_odoo_message(odoo: dict, task_id: int, body: str):
    odoo_create(odoo, "mail.message", {
        "model": "x_auto_posting", "res_id": task_id,
        "message_type": "comment", "subtype_id": 1, "body": body,
    })


def set_task_active(odoo: dict, task_ids: list[int]):
    if not task_ids:
        return
    odoo_write(odoo, "x_auto_posting", task_ids, {
        "x_studio_stage_id": STAGE_ACTIVE, "x_studio_add_from_api": True, "x_studio_url_listing": False,
    })
    for tid in task_ids:
        send_odoo_message(odoo, tid, 'Товар успешно залищен со статусом "LIVE" market: Worten PT | HAJUS')


def set_task_new(odoo: dict, task_ids: list[int]):
    if not task_ids:
        return
    odoo_write(odoo, "x_auto_posting", task_ids, {"x_studio_url_listing": False, "x_studio_stage_id": STAGE_NEW_TASK})
    for tid in task_ids:
        send_odoo_message(odoo, tid, "Товар не залищен на Worten PT | HAJUS, повторно отправлен в New task Stage")


def build_mcm_indexes(mcm_dump: list[dict]) -> tuple[dict, dict]:
    by_pid: dict[str, dict] = {}
    by_ean: dict[str, dict] = {}
    for it in mcm_dump:
        pid = it.get("provider_unique_identifier")
        if pid:
            by_pid[str(pid).strip()] = it
        for uid in (it.get("unique_identifiers") or []):
            if uid.get("code") == "EAN" and uid.get("value"):
                by_ean[str(uid["value"]).strip()] = it
    return by_pid, by_ean


def classify_products_v3(product_ids: list[int], products_data: dict, product_map: dict,
                          mcm_by_pid: dict, mcm_by_ean: dict) -> dict:
    """v3: LIVE -> always blind-import (no offer-check). NOT_LIVE/missing as before."""
    results = {
        "new_task": [], "not_live": [],
        "live_pids": [], "pid_to_sku": {},
        "status_breakdown": {},
    }

    for pid in product_ids:
        product = products_data.get(pid)
        if not product:
            continue
        task_ids = product_map[pid]
        sku = clean_sku(product.get("default_code", "") or "")
        ean = (product.get("barcode") or "").strip()

        item = None
        if sku and sku in mcm_by_pid:
            item = mcm_by_pid[sku]
        elif ean and ean in mcm_by_ean:
            item = mcm_by_ean[ean]

        if not item:
            results["new_task"].extend(task_ids)
            results["status_breakdown"]["NOT_IN_MCM"] = results["status_breakdown"].get("NOT_IN_MCM", 0) + 1
            continue

        status = item.get("status")
        if status == "LIVE":
            results["live_pids"].append(pid)
            results["pid_to_sku"][pid] = sku
            results["status_breakdown"]["LIVE"] = results["status_breakdown"].get("LIVE", 0) + 1
        else:
            results["not_live"].extend(task_ids)
            err_codes = ",".join(sorted({(e.get("code") or "") for e in (item.get("errors") or [])})) or "-"
            key = f"NOT_LIVE|{err_codes}"
            results["status_breakdown"][key] = results["status_breakdown"].get(key, 0) + 1

    return results


def _resources_from_env() -> tuple[dict, dict]:
    odoo = {
        "url": os.environ["ODOO_URL"],
        "db": os.environ["ODOO_DB"],
        "uid": int(os.environ["ODOO_UID"]),
        "api_key": os.environ["ODOO_API_KEY"],
    }
    worten = {
        "base_url": os.environ["WORTEN_BASE_URL"],
        "api_key": os.environ["WORTEN_API_KEY"],
        "shop_id": int(os.environ["WORTEN_SHOP_ID"]),
    }
    return odoo, worten


def main():
    batch_limit = int(os.environ.get("BATCH_LIMIT", "0"))
    try:
        odoo, worten = _resources_from_env()
    except KeyError as e:
        print(f"Missing required environment variable: {e}", file=sys.stderr)
        sys.exit(2)
    try:
        return _run(odoo, worten, batch_limit)
    except Exception as e:
        send_alert(str(e), traceback.format_exc())
        raise


def _run(odoo: dict, worten: dict, batch_limit: int = 0):
    t0 = time.time()
    print("Fetching pending tasks from Odoo (Worten website 117)...")
    tasks = odoo_rpc(odoo, "x_auto_posting", "search_read", [
        ["x_studio_stage_id", "=", STAGE_PENDING],
        ["x_studio_website", "in", PENDING_WEBSITE_IDS],
    ], fields=["id", "x_studio_website", "x_studio_product", "x_studio_stage_id"])

    if not tasks:
        print("No pending tasks found.")
        return {"processed": 0, "active": 0, "new_task": 0, "not_live": 0, "offers_created": 0}

    print(f"  Found {len(tasks)} pending tasks")

    product_map: dict[int, list[int]] = {}
    for t in tasks:
        prod = t.get("x_studio_product")
        if not prod:
            continue
        pid = prod[0] if isinstance(prod, list) else prod
        product_map.setdefault(pid, []).append(t["id"])

    product_ids = list(product_map.keys())
    if batch_limit:
        product_ids = product_ids[:batch_limit]

    print(f"  Unique products: {len(product_ids)}")

    # --- MCM dump first (cheap) ---
    t_mcm = time.time()
    print("Fetching whole-shop MCM status dump (1 request)...")
    mcm_dump = mirakl_fetch_mcm_status_dump(worten)
    print(f"  MCM dump: {len(mcm_dump)} items in {time.time() - t_mcm:.1f}s")
    mcm_by_pid, mcm_by_ean = build_mcm_indexes(mcm_dump)

    # --- Preliminary classification by SKU/EAN against MCM (no product data yet) ---
    print("Pre-classifying products from MCM by SKU/EAN...")
    print("Fetching minimal product identifiers from Odoo...")
    products_min: dict[int, dict] = {}
    CHUNK = 200
    for i in range(0, len(product_ids), CHUNK):
        chunk = product_ids[i:i + CHUNK]
        records = odoo_rpc(odoo, "product.template", "search_read", [["id", "in", chunk]],
                           fields=["id", "barcode", "default_code"])
        for r in records:
            products_min[r["id"]] = r

    prelim = classify_products_v3(product_ids, products_min, product_map, mcm_by_pid, mcm_by_ean)
    live_pids = prelim["live_pids"]
    to_new_task = list(prelim["new_task"])
    to_not_live = list(prelim["not_live"])
    breakdown = prelim["status_breakdown"]

    print(f"\nPre-classification results:")
    print(f"  LIVE products to blind-import: {len(live_pids)}")
    print(f"  NOT_LIVE tasks: {len(to_not_live)}")
    print(f"  NOT_IN_MCM (->New Task): {len(to_new_task)}")
    print(f"  Status breakdown:")
    for k, v in sorted(breakdown.items(), key=lambda x: -x[1]):
        print(f"    {k}: {v}")

    to_activate: list[int] = []
    activated_pids: set[int] = set()
    offers_created = 0
    offers_failed = 0

    # --- Build offers only for LIVE products ---
    if live_pids:
        print(f"\nFetching full product data for {len(live_pids)} LIVE products...")
        products_data: dict[int, dict] = {}
        for i in range(0, len(live_pids), CHUNK):
            chunk = live_pids[i:i + CHUNK]
            records = odoo_rpc(odoo, "product.template", "search_read", [["id", "in", chunk]], fields=[
                "id", "barcode", "default_code", "list_price", "taxes_id",
                "x_studio_discount", "dr_free_qty",
            ])
            for r in records:
                products_data[r["id"]] = r

        print("Fetching pricelist data (PT=20, ES=24) for LIVE products...")
        # prices_by_product[product_template_id][pricelist_id] = fixed_price
        prices_by_product: dict[int, dict[int, float]] = {pid: {} for pid in live_pids}
        for i in range(0, len(live_pids), CHUNK):
            chunk = live_pids[i:i + CHUNK]
            prices = odoo_rpc(odoo, "product.pricelist.item", "search_read", [
                ["pricelist_id", "in", list(PRICELIST_TO_CHANNEL.keys())],
                ["product_tmpl_id", "in", chunk],
                ["compute_price", "=", "fixed"],
            ], fields=["fixed_price", "product_tmpl_id", "pricelist_id"])
            for p in prices:
                tmpl = p.get("product_tmpl_id")
                if isinstance(tmpl, list):
                    tmpl = tmpl[0]
                plist = p.get("pricelist_id")
                if isinstance(plist, list):
                    plist = plist[0]
                fp = p.get("fixed_price", 0) or 0
                if fp > 0 and tmpl in prices_by_product and plist in PRICELIST_TO_CHANNEL:
                    prices_by_product[tmpl][plist] = fp

        print(f"\nBuilding offers for {len(live_pids)} LIVE products...")
        offer_rows = []
        skipped_no_price: list[str] = []
        for pid in live_pids:
            product = products_data.get(pid)
            if not product:
                continue
            row = build_offer_row(product, prices_by_product.get(pid, {}))
            if row is None:
                skipped_no_price.append(clean_sku(product.get("default_code", "") or ""))
                to_new_task.extend(product_map.get(pid, []))
                continue
            offer_rows.append(row)

        if skipped_no_price:
            print(f"  Skipped {len(skipped_no_price)} LIVE products with no PT/ES pricelist data -> New Task")

        if offer_rows:
            print(f"\nUploading {len(offer_rows)} offers as XLSX (blind-import)...")
            xlsx_data = build_xlsx(offer_rows)
            import_id = mirakl_import_offers(worten, xlsx_data)
            print(f"  Import ID: {import_id}")
            for attempt in range(360):
                time.sleep(10)
                status = mirakl_check_import(worten, import_id)
                print(f"  Import status: {status.get('status', 'unknown')} (attempt {attempt + 1})")
                if status.get("status") in ("COMPLETE", "WAITING"):
                    lines_ok = status.get("lines_in_success", 0)
                    lines_err = status.get("lines_in_error", 0)
                    offers_created = status.get("offer_inserted", 0) + status.get("offer_updated", 0)
                    print(f"  Import result: {lines_ok} ok, {lines_err} errors, {offers_created} inserted/updated")

                    failed_skus: set[str] = set()
                    if lines_err > 0 and status.get("has_error_report"):
                        try:
                            failed_skus = mirakl_get_failed_skus(worten, import_id)
                            print(f"  Failed SKUs ({len(failed_skus)})")
                        except Exception as e:
                            print(f"  Could not fetch error report: {e}")

                    for pid in live_pids:
                        sku = prelim["pid_to_sku"].get(pid, "")
                        if sku in skipped_no_price:
                            continue
                        if sku in failed_skus:
                            offers_failed += 1
                            to_new_task.extend(product_map.get(pid, []))
                        else:
                            to_activate.extend(product_map.get(pid, []))
                            activated_pids.add(pid)

                    break
            else:
                print("  Import timed out after 60 minutes")

    # --- Cross-website fan-out for accepted products ---
    # Worten has a single Odoo website (117). Kept as a defensive sweep in case
    # Odoo grows a second Worten website later.
    if activated_pids and len(WEBSITE_IDS) > 1:
        print(f"\nFetching cross-website tasks for {len(activated_pids)} accepted products...")
        original_count = len(to_activate)
        activate_ids = set(to_activate)
        new_task_ids = set(to_new_task)
        activated_pid_list = list(activated_pids)
        for i in range(0, len(activated_pid_list), CHUNK):
            chunk = activated_pid_list[i:i + CHUNK]
            extra_tasks = odoo_rpc(odoo, "x_auto_posting", "search_read", [
                ["x_studio_product", "in", chunk],
                ["x_studio_website", "in", WEBSITE_IDS],
                ["x_studio_stage_id", "!=", STAGE_ACTIVE],
            ], fields=["id"])
            for t in extra_tasks:
                if t["id"] not in new_task_ids:
                    activate_ids.add(t["id"])
        to_activate = list(activate_ids)
        print(f"  Tasks to activate: {len(to_activate)} (was {original_count}, +{len(to_activate) - original_count} from cross-website)")

    # --- Odoo updates ---
    print("\nUpdating Odoo tasks...")
    if to_activate:
        print(f"  Setting {len(to_activate)} tasks to Active...")
        set_task_active(odoo, to_activate)
    if to_new_task:
        print(f"  Setting {len(to_new_task)} tasks to New Task...")
        set_task_new(odoo, to_new_task)
    if to_not_live:
        print(f"  Leaving {len(to_not_live)} NOT_LIVE tasks in Pending (no per-task comment)")

    elapsed = time.time() - t0
    result = {
        "total_tasks": len(tasks),
        "unique_products": len(product_ids),
        "live_products": len(live_pids),
        "activated": len(to_activate),
        "new_task": len(to_new_task),
        "not_live": len(to_not_live),
        "offers_created": offers_created,
        "offers_failed": offers_failed,
        "elapsed_sec": round(elapsed, 1),
        "mcm_items": len(mcm_dump),
    }
    print(f"\nDone in {elapsed:.1f}s! {result}")
    return result


if __name__ == "__main__":
    main()
